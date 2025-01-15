import os
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as XlImage
import json

from .share import *


#############
# MARK: Ruler
#############
def render_ruler(document, ws):
    """定規の描画
    """
    print("🔧　定規の描画")

    HORIZONTAL_RULER_HEIGHT = 2     # 垂直定規の縦幅
    VERTICAL_RULER_WIDTH = 2        # 垂直定規の横幅

    # Trellis では、タテ：ヨコ＝３：３ で、１ユニットセルとします。
    # また、上辺、右辺、下辺、左辺に、１セル幅の定規を置きます
    canvas_rect = Rectangle.from_dict(document['canvas'])

    # 横幅または縦幅が１アウト未満の場合は、定規は描画しません
    if canvas_rect.width_obj.total_of_out_counts_qty < 1 or canvas_rect.height_obj.total_of_out_counts_qty < 1:
        return

    # 行の横幅
    for column_th in range(
            canvas_rect.left_obj.total_of_out_counts_th,
            canvas_rect.left_obj.total_of_out_counts_th + canvas_rect.width_obj.total_of_out_counts_qty):
        column_letter = xl.utils.get_column_letter(column_th)
        ws.column_dimensions[column_letter].width = 2.7    # 2.7 characters = about 30 pixels

    # 列の高さ
    for row_th in range(
            canvas_rect.top_obj.total_of_out_counts_th,
            canvas_rect.top_obj.total_of_out_counts_th + canvas_rect.height_obj.total_of_out_counts_qty):
        ws.row_dimensions[row_th].height = 15    # 15 points = about 30 pixels


    # 定規を描画しないケース
    if (
            # ruler 項目がない、 
            'ruler' not in document or
            # ruler 項目にヌルが設定されている
            (ruler_dict := document['ruler']) is None or
            # document.visibule プロパティがない
            'visible' not in ruler_dict or
            # document.visibule プロパティがヌルか偽だ
            ruler_dict['visible'] in [None, False]):
        return


    # ウィンドウ枠の固定
    ws.freeze_panes = 'C2'

    # 定規の文字色。２色固定
    # TODO 任意色数に対応したい
    if 'fgColor' in ruler_dict and (fg_color_list := ruler_dict['fgColor']) is not None:
        
        if (fg_color_text := fg_color_list[0]) == 'paper_color':
            #first_font = Font(color=None)   # フォントに使うと黒になる
            raise ValueError('fgColor に paper_color を指定してはいけません')

        elif (first_font_text := tone_and_color_name_to_color_code(fg_color_text)) and first_font_text is not None:
            try:
                first_font = Font(color=first_font_text)
            except:
                print(f'ERROR: {first_font_text=}')
                raise

        if (fg_color_text := fg_color_list[1]) == 'paper_color':
            #second_font = Font(color=None)   # フォントに使うと黒になる
            raise ValueError('fgColor に paper_color を指定してはいけません')

        elif (second_font_text := tone_and_color_name_to_color_code(fg_color_text)) and second_font_text is not None:
            try:
                second_font = Font(color=second_font_text)
            except:
                print(f'ERROR: {second_font_text=}')
                raise

    else:
        # フォントの色の既定値は黒
        first_font = Font(color='000000')
        second_font = Font(color='000000')

    # 定規の背景色。２色固定
    # TODO 任意色数に対応したい
    if 'bgColor' in ruler_dict and (bg_color_list := ruler_dict['bgColor']) is not None:
        
        if (bg_color_text := bg_color_list[0]) == 'paper_color':
            first_pattern_fill = PatternFill(patternType=None)

        elif (first_pattern_fill_text := tone_and_color_name_to_color_code(bg_color_text)) and first_pattern_fill_text is not None:
            try:
                first_pattern_fill = PatternFill(patternType='solid', fgColor=first_pattern_fill_text)
            except:
                print(f'ERROR: {first_pattern_fill_text=}')
                raise

        if (bg_color_text := bg_color_list[1]) == 'paper_color':
            second_pattern_fill = PatternFill(patternType=None)

        elif (second_pattern_fill_text := tone_and_color_name_to_color_code(bg_color_text)) and second_pattern_fill_text is not None:
            try:
                second_pattern_fill = PatternFill(patternType='solid', fgColor=second_pattern_fill_text)
            except:
                print(f'ERROR: {second_pattern_fill_text=}')
                raise

    else:
        # フォントの色の既定値は白
        first_pattern_fill = PatternFill(patternType='solid', fgColor='FFFFFF')
        second_pattern_fill = PatternFill(patternType='solid', fgColor='FFFFFF')

    center_center_alignment = Alignment(horizontal='center', vertical='center')


    def render_ruler_numbering_and_coloring_of_top_edge():
        """定規の採番と着色　＞　上辺

                横幅が３で割り切れるとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 3 シュリンクする
                ■■□[  1 ][  2 ]□■■
                ■■                ■■

                横幅が３で割ると１余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 4 シュリンクする
                ■■□[  1 ][  2 ]□□■■
                ■■                  ■■

                横幅が３で割ると２余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 2 シュリンクする
                ■■□[  1 ][  2 ][  3 ]■■
                ■■                    ■■
        """
        row_th = canvas_rect.top_obj.total_of_out_counts_th

        for column_th in range(
                canvas_rect.left_obj.total_of_out_counts_th + OUT_COUNTS_THAT_CHANGE_INNING,
                canvas_rect.right_obj.total_of_out_counts_th - OUT_COUNTS_THAT_CHANGE_INNING,
                OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{row_th}']

            # 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12,
            # -------- -------- -------- -----------
            # dark      light    dark     light
            #
            # - 1 する
            #
            # 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11,
            # -------- -------- -------- ----------
            # dark     light    dark     light
            #
            # 3 で割って端数を切り捨て
            #
            # 0, 0, 0, 1, 1, 1, 2, 2, 2, 3, 3, 3,
            # -------- -------- -------- --------
            # dark     light    dark     light
            #
            # 2 で割った余り
            #
            # 0, 0, 0, 1, 1, 1, 0, 0, 0, 1, 1, 1,
            # -------- -------- -------- --------
            # dark     light    dark     light
            #
            ruler_number = (column_th - canvas_rect.left_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
            is_left_end = (column_th - canvas_rect.left_obj.total_of_out_counts_th) % OUT_COUNTS_THAT_CHANGE_INNING == 0

            if is_left_end:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                if ruler_number % 2 == 0:
                    cell.font = first_font
                else:
                    cell.font = second_font

            if ruler_number % 2 == 0:
                cell.fill = first_pattern_fill
            else:
                cell.fill = second_pattern_fill


    def render_ruler_numbering_and_coloring_of_left_edge():
        """定規の採番と着色　＞　左辺

        縦幅が３で割り切れるとき、１投球回は 1th から始まる。最後の投球回は、端数なしで表示できる
        [  0 ][  1 ][  2 ][  3 ]
        ■                    ■

        縦幅が３で割ると１余るとき、１投球回は 1th から始まる。最後の投球回は、端数１になる
        [  0 ][  1 ][  2 ][  3 ]□
        ■                      ■

        縦幅が３で割ると２余るとき、１投球回は 1th から始まる。最後の投球回は、端数２になる
        [  0 ][  1 ][  2 ][  3 ]□□
        ■                        ■
        """

        # 幅が４アウト未満の場合、左辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_rect.width_obj.total_of_out_counts_qty < 4:
            return

        column_th = canvas_rect.left_obj.total_of_out_counts_th
        column_letter = xl.utils.get_column_letter(column_th)
        shrink = canvas_rect.height_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING

        for row_th in range(
                canvas_rect.top_obj.total_of_out_counts_th,
                canvas_rect.bottom_obj.total_of_out_counts_th - shrink,
                OUT_COUNTS_THAT_CHANGE_INNING):
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
            is_top_end = (row_th - canvas_rect.top_obj.total_of_out_counts_th) % OUT_COUNTS_THAT_CHANGE_INNING == 0

            if is_top_end:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                if ruler_number % 2 == 0:
                    cell.font = first_font
                else:
                    cell.font = second_font

            if ruler_number % 2 == 0:
                cell.fill = first_pattern_fill
            else:
                cell.fill = second_pattern_fill


    def render_ruler_coloring_of_left_edge_bottom_spacing():
        """左辺の最後の要素が端数のとき、左辺の最後の要素の左上へ着色

                最後の端数の要素に色を塗ってもらいたいから、もう１要素着色しておく
        """
        vertical_remain = canvas_rect.height_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        #print(f'左辺 h_qty={canvas_rect.height_obj.total_of_out_counts_qty} {shrink=} {vertical_remain=}')

        if vertical_remain != 0:
            column_th = canvas_rect.left_obj.total_of_out_counts_th
            column_letter = xl.utils.get_column_letter(column_th)
            row_th = canvas_rect.bottom_obj.total_of_out_counts_th - vertical_remain
            ruler_number = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
            #print(f"""左辺の最後の要素の左上へ着色 {row_th=} {ruler_number=}""")
            cell = ws[f'{column_letter}{row_th}']

            # 数字も振りたい
            if vertical_remain == 2:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                if ruler_number % 2 == 0:
                    cell.font = first_font
                else:
                    cell.font = second_font

            if ruler_number % 2 == 0:
                cell.fill = first_pattern_fill
            else:
                cell.fill = second_pattern_fill


    def render_ruler_numbering_and_coloring_of_bottom_edge():
        """定規の採番と着色　＞　下辺
        """

        # 高さが２投球回未満の場合、下辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_rect.height_obj.total_of_out_counts_qty < 2:
            return

        row_th = canvas_rect.bottom_obj.total_of_out_counts_th - 1
        bottom_is_first_pattern_fill = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING % 2 == 0

        for column_th in range(
                canvas_rect.left_obj.total_of_out_counts_th + OUT_COUNTS_THAT_CHANGE_INNING,
                canvas_rect.right_obj.total_of_out_counts_th - OUT_COUNTS_THAT_CHANGE_INNING,
                OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{row_th}']
            ruler_number = (column_th - canvas_rect.left_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
            is_left_end = (column_th - canvas_rect.left_obj.total_of_out_counts_th) % OUT_COUNTS_THAT_CHANGE_INNING == 0

            if is_left_end:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                if ruler_number % 2 == 0:
                    if bottom_is_first_pattern_fill:
                        cell.font = first_font
                    else:
                        cell.font = second_font
                else:
                    if bottom_is_first_pattern_fill:
                        cell.font = second_font
                    else:
                        cell.font = first_font

            if ruler_number % 2 == 0:
                if bottom_is_first_pattern_fill:
                    cell.fill = first_pattern_fill
                else:
                    cell.fill = second_pattern_fill
            else:
                if bottom_is_first_pattern_fill:
                    cell.fill = second_pattern_fill
                else:
                    cell.fill = first_pattern_fill


    def render_ruler_numbering_and_coloring_of_right_edge():
        """定規の採番と着色　＞　右辺
        """

        # 幅が４アウト未満の場合、右辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_rect.width_obj.total_of_out_counts_qty < 4:
            return

        column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH
        column_letter = xl.utils.get_column_letter(column_th)
        rightest_is_first_pattern_fill = (column_th - canvas_rect.left_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING % 2 == 0
        shrink = canvas_rect.height_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING

        for row_th in range(
                canvas_rect.top_obj.total_of_out_counts_th,
                canvas_rect.bottom_obj.total_of_out_counts_th - shrink,
                OUT_COUNTS_THAT_CHANGE_INNING):
            cell = ws[f'{column_letter}{row_th}']

            ruler_number = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
            is_top_end = (row_th - canvas_rect.top_obj.total_of_out_counts_th) % OUT_COUNTS_THAT_CHANGE_INNING == 0

            if is_top_end:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                if ruler_number % 2 == 0:
                    if rightest_is_first_pattern_fill:
                        cell.font = first_font
                    else:
                        cell.font = second_font
                else:
                    if rightest_is_first_pattern_fill:
                        cell.font = second_font
                    else:
                        cell.font = first_font

            if ruler_number % 2 == 0:
                if rightest_is_first_pattern_fill:
                    cell.fill = first_pattern_fill
                else:
                    cell.fill = second_pattern_fill
            else:
                if rightest_is_first_pattern_fill:
                    cell.fill = second_pattern_fill
                else:
                    cell.fill = first_pattern_fill


    def render_ruler_coloring_of_right_edge_bottom_spacing():
        """右辺の最後の要素が端数のとき、右辺の最後の要素の左上へ着色

                最後の端数の要素に色を塗ってもらいたいから、もう１要素着色しておく
        """
        vertical_remain = canvas_rect.height_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        #print(f'右辺 h_qty={canvas_rect.height_obj.total_of_out_counts_qty} {shrink=} {vertical_remain=}')

        if vertical_remain != 0:
            column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH
            column_letter = xl.utils.get_column_letter(column_th)
            row_th = canvas_rect.bottom_obj.total_of_out_counts_th - vertical_remain
            ruler_number = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
            #print(f"""右辺の最後の要素の左上へ着色 {row_th=} {ruler_number=}""")
            cell = ws[f'{column_letter}{row_th}']

            rightest_is_first_pattern_fill = (column_th - canvas_rect.left_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING % 2 == 0

            # 数字も振りたい
            if vertical_remain == 2:
                cell.value = ruler_number
                cell.alignment = center_center_alignment
                if ruler_number % 2 == 0:
                    if rightest_is_first_pattern_fill:
                        cell.font = first_font
                    else:
                        cell.font = second_font
                else:
                    if rightest_is_first_pattern_fill:
                        cell.font = second_font
                    else:
                        cell.font = first_font

            if ruler_number % 2 == 0:
                if rightest_is_first_pattern_fill:
                    cell.fill = first_pattern_fill
                else:
                    cell.fill = second_pattern_fill
            else:
                if rightest_is_first_pattern_fill:
                    cell.fill = second_pattern_fill
                else:
                    cell.fill = first_pattern_fill


    def render_ruler_coloring_of_top_left_spacing():
        """定規の着色　＞　左上の１セルの隙間
        """
        column_th = canvas_rect.left_obj.total_of_out_counts_th + VERTICAL_RULER_WIDTH
        row_th = canvas_rect.top_obj.total_of_out_counts_th
        ruler_number = (column_th - canvas_rect.left_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        if ruler_number % 2 == 0:
            cell.fill = first_pattern_fill
        else:
            cell.fill = second_pattern_fill


    def render_ruler_coloring_right_end_spacing_on_top():
        """定規の着色　＞　上の水平定規の右端の隙間の先頭
        """
        horizontal_remain = canvas_rect.width_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain in [1, 2]:
            return

        row_th = canvas_rect.top_obj.total_of_out_counts_th

        # 何アウト余るか
        spacing = (canvas_rect.width_obj.total_of_out_counts_qty - VERTICAL_RULER_WIDTH) % OUT_COUNTS_THAT_CHANGE_INNING

        # 隙間の先頭
        column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH - spacing
        column_letter = xl.utils.get_column_letter(column_th)

        # 隙間に表示される定規の番号
        ruler_number = column_th // OUT_COUNTS_THAT_CHANGE_INNING

        cell = ws[f'{column_letter}{row_th}']
        if ruler_number % 2 == 0:
            cell.fill = first_pattern_fill
        else:
            cell.fill = second_pattern_fill


    def render_ruler_coloring_of_bottom_left_spacing():
        """定規の着色　＞　左下の１セルの隙間
        """
        column_th = canvas_rect.left_obj.total_of_out_counts_th + VERTICAL_RULER_WIDTH
        row_th = canvas_rect.bottom_obj.total_of_out_counts_th - 1
        bottom_is_first_pattern_fill = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING % 2 == 0

        ruler_number = (column_th - canvas_rect.left_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        if ruler_number % 2 == 0:
            if bottom_is_first_pattern_fill:
                cell.fill = first_pattern_fill
            else:
                cell.fill = second_pattern_fill
        else:
            if bottom_is_first_pattern_fill:
                cell.fill = second_pattern_fill
            else:
                cell.fill = first_pattern_fill


    def render_ruler_coloring_right_end_spacing_on_bottom():
        """定規の着色　＞　下の水平定規の右端の隙間の先頭
        """
        horizontal_remain = canvas_rect.width_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain in [1, 2]:
            return

        row_th = canvas_rect.bottom_obj.total_of_out_counts_th - 1
        bottom_is_first_pattern_fill = (row_th - canvas_rect.top_obj.total_of_out_counts_th) // OUT_COUNTS_THAT_CHANGE_INNING % 2 == 0

        # 何アウト余るか
        spacing = (canvas_rect.width_obj.total_of_out_counts_qty - VERTICAL_RULER_WIDTH) % OUT_COUNTS_THAT_CHANGE_INNING

        # 隙間の先頭
        column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH - spacing
        column_letter = xl.utils.get_column_letter(column_th)

        # 隙間に表示される定規の番号
        ruler_number = column_th // OUT_COUNTS_THAT_CHANGE_INNING

        cell = ws[f'{column_letter}{row_th}']
        if ruler_number % 2 == 0:
            if bottom_is_first_pattern_fill:
                cell.fill = first_pattern_fill
            else:
                cell.fill = second_pattern_fill
        else:
            if bottom_is_first_pattern_fill:
                cell.fill = second_pattern_fill
            else:
                cell.fill = first_pattern_fill


    def render_ruler_merge_cells_of_top_edge():
        """定規のセル結合　＞　上辺

        横幅が３で割り切れるとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 3 シュリンクする
        ■■□[  1 ][  2 ]□■■
        ■■                ■■

        横幅が３で割ると１余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 4 シュリンクする
        ■■□[  1 ][  2 ]□□■■
        ■■                  ■■

        横幅が３で割ると２余るとき、１投球回は 4th から始まる。２投球回を最終表示にするためには、横幅を 2 シュリンクする
        ■■□[  1 ][  2 ][  3 ]■■
        ■■                    ■■
        """
        skip_left = OUT_COUNTS_THAT_CHANGE_INNING
        horizontal_remain = canvas_rect.width_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain == 0:
            shrink_right = 3
        elif horizontal_remain == 1:
            shrink_right = 4
        else:
            shrink_right = 2

        row_th = canvas_rect.top_obj.total_of_out_counts_th

        for column_th in range(
                canvas_rect.left_obj.total_of_out_counts_th + skip_left,
                canvas_rect.right_obj.total_of_out_counts_th - shrink_right,
                OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + 2)
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    def render_ruler_merge_cells_of_left_edge():
        """定規のセル結合　＞　左辺

        縦幅が３で割り切れるとき、１投球回は 1th から始まる。最後の投球回は、端数なしで表示できる
        [  0 ][  1 ][  2 ][  3 ]
        ■                    ■

        縦幅が３で割ると１余るとき、１投球回は 1th から始まる。最後の投球回は、端数１になる
        [  0 ][  1 ][  2 ][  3 ]□
        ■                      ■

        縦幅が３で割ると２余るとき、１投球回は 1th から始まる。最後の投球回は、端数２になる
        [  0 ][  1 ][  2 ][  3 ]□□
        ■                        ■
        """

        # 幅が４アウト未満の場合、左辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_rect.width_obj.total_of_out_counts_qty < 4:
            return

        column_th = canvas_rect.left_obj.total_of_out_counts_th
        column_letter = xl.utils.get_column_letter(column_th)
        column_letter2 = xl.utils.get_column_letter(column_th + 1)

        for row_th in range(
                canvas_rect.top_obj.total_of_out_counts_th,
                canvas_rect.bottom_obj.total_of_out_counts_th - OUT_COUNTS_THAT_CHANGE_INNING,
                OUT_COUNTS_THAT_CHANGE_INNING):
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')

        # 高さが１イニング未満の場合、最後の要素はありません
        if canvas_rect.height_obj.total_of_out_counts_qty < OUT_COUNTS_THAT_CHANGE_INNING:
            return
        
        # 最後の要素
        spacing = canvas_rect.height_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 0:
            row_th = canvas_rect.height_obj.integer_part * OUT_COUNTS_THAT_CHANGE_INNING + canvas_rect.top_obj.total_of_out_counts_th - OUT_COUNTS_THAT_CHANGE_INNING
            #print(f'マージセルA h_qty={canvas_rect.height_obj.total_of_out_counts_qty} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')
        elif spacing == 1:
            row_th = canvas_rect.height_obj.integer_part * OUT_COUNTS_THAT_CHANGE_INNING + canvas_rect.top_obj.total_of_out_counts_th
            #print(f'マージセルB {row_th=} {spacing=} {column_letter=} {column_letter2=} {canvas_rect.height_obj.integer_part=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')
        elif spacing == 2:
            row_th = canvas_rect.height_obj.integer_part * OUT_COUNTS_THAT_CHANGE_INNING + canvas_rect.top_obj.total_of_out_counts_th
            #print(f'マージセルH h_qty={canvas_rect.height_obj.total_of_out_counts_qty} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 1}')


    def render_ruler_merge_cells_of_bottom_edge():
        """定規のセル結合　＞　下辺"""

        # 高さが２投球回未満の場合、下辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_rect.height_obj.total_of_out_counts_qty < 2:
            return

        skip_left = OUT_COUNTS_THAT_CHANGE_INNING
        horizontal_remain = canvas_rect.width_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        if horizontal_remain == 0:
            shrink_right = 3
        elif horizontal_remain == 1:
            shrink_right = 4
        else:
            shrink_right = 2

        row_th = canvas_rect.bottom_obj.total_of_out_counts_th - 1

        for column_th in range(
                canvas_rect.left_obj.total_of_out_counts_th + skip_left,
                canvas_rect.right_obj.total_of_out_counts_th - shrink_right,
                OUT_COUNTS_THAT_CHANGE_INNING):
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + 2)
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    def render_ruler_merge_cells_of_right_edge():
        """定規のセル結合　＞　右辺"""

        # 幅が４アウト未満の場合、右辺のルーラーは描かないものとします（上、右、下、左の辺の定規のセル結合が被ってしまうため、上辺だけ残します）
        if canvas_rect.width_obj.total_of_out_counts_qty < 4:
            return

        column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH
        column_letter = xl.utils.get_column_letter(column_th)
        column_letter2 = xl.utils.get_column_letter(column_th + 1)

        for row_th in range(
                canvas_rect.top_obj.total_of_out_counts_th,
                canvas_rect.bottom_obj.total_of_out_counts_th - OUT_COUNTS_THAT_CHANGE_INNING,
                OUT_COUNTS_THAT_CHANGE_INNING):
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')

        # 高さが１イニング未満の場合、最後の要素はありません
        if canvas_rect.height_obj.total_of_out_counts_qty < OUT_COUNTS_THAT_CHANGE_INNING:
            return
        
        # 最後の要素
        spacing = canvas_rect.height_obj.total_of_out_counts_qty % OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 0:
            row_th = canvas_rect.height_obj.integer_part * OUT_COUNTS_THAT_CHANGE_INNING + canvas_rect.top_obj.total_of_out_counts_th - OUT_COUNTS_THAT_CHANGE_INNING
            #print(f'マージセルC h_qty={canvas_rect.height_obj.total_of_out_counts_qty} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 2}')
        elif spacing == 1:
            row_th = canvas_rect.height_obj.integer_part * OUT_COUNTS_THAT_CHANGE_INNING + canvas_rect.top_obj.total_of_out_counts_th
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')
        elif spacing == 2:
            row_th = canvas_rect.height_obj.integer_part * OUT_COUNTS_THAT_CHANGE_INNING + canvas_rect.top_obj.total_of_out_counts_th
            #print(f'マージセルD h_qty={canvas_rect.height_obj.total_of_out_counts_qty} {row_th=} {spacing=}')
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + 1}')


    def render_ruler_merge_cells_right_end_fraction_on_top():
        """上側の水平［定規］の右端の端数のセル結合"""

        # 隙間の幅
        spacing = (canvas_rect.width_obj.total_of_out_counts_qty - VERTICAL_RULER_WIDTH) % OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 2:
            column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH - spacing
            row_th = canvas_rect.top_obj.total_of_out_counts_th
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + spacing - 1)
            #print(f"""マージセルE {column_th=} {row_th=} {column_letter=} {column_letter2=}""")
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    def render_ruler_merge_cells_right_end_fraction_on_bottom():
        """下側の水平［定規］の右端の端数のセル結合"""

        # 隙間の幅
        spacing = (canvas_rect.width_obj.total_of_out_counts_qty - VERTICAL_RULER_WIDTH) % OUT_COUNTS_THAT_CHANGE_INNING
        if spacing == 2:
            column_th = canvas_rect.right_obj.total_of_out_counts_th - VERTICAL_RULER_WIDTH - spacing
            row_th = canvas_rect.bottom_obj.total_of_out_counts_th - 1
            column_letter = xl.utils.get_column_letter(column_th)
            column_letter2 = xl.utils.get_column_letter(column_th + spacing - 1)
            #print(f"""マージセルF {column_th=} {row_th=} {column_letter=} {column_letter2=}""")
            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th}')


    # 定規の採番と着色　＞　上辺
    render_ruler_numbering_and_coloring_of_top_edge()

    # 定規の採番と着色　＞　左辺
    render_ruler_numbering_and_coloring_of_left_edge()

    # 定規の採番と着色　＞　下辺
    render_ruler_numbering_and_coloring_of_bottom_edge()

    # 定規の採番と着色　＞　右辺
    render_ruler_numbering_and_coloring_of_right_edge()

    # 左辺の最後の要素が端数のとき、左辺の最後の要素の左上へ着色
    render_ruler_coloring_of_left_edge_bottom_spacing()

    # 右辺の最後の要素が端数のとき、右辺の最後の要素の左上へ着色
    render_ruler_coloring_of_right_edge_bottom_spacing()

    # NOTE 上下の辺の両端の端数の処理

    # 定規の着色　＞　左上の１セルの隙間
    render_ruler_coloring_of_top_left_spacing()

    # 定規の着色　＞　上の水平定規の右端の隙間の先頭
    render_ruler_coloring_right_end_spacing_on_top()

    # 定規の着色　＞　左下の１セルの隙間
    render_ruler_coloring_of_bottom_left_spacing()

    # 定規の着色　＞　下の水平定規の右端の隙間の先頭
    render_ruler_coloring_right_end_spacing_on_bottom()

    # NOTE セル結合すると read only セルになるから、セル結合は、セルを編集が終わったあとで行う

    # 定規のセル結合　＞　上辺
    render_ruler_merge_cells_of_top_edge()

    # 定規のセル結合　＞　左辺
    render_ruler_merge_cells_of_left_edge()

    # 定規のセル結合　＞　下辺
    render_ruler_merge_cells_of_bottom_edge()

    # 定規のセル結合　＞　右辺
    render_ruler_merge_cells_of_right_edge()

    # 上側の水平［定規］の右端の端数のセル結合
    render_ruler_merge_cells_right_end_fraction_on_top()

    # 下側の水平［定規］の右端の端数のセル結合
    render_ruler_merge_cells_right_end_fraction_on_bottom()
