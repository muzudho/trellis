import openpyxl as xl
import os

from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

from ..share import Canvas, Card, ColorSystem, Pillar, Rectangle, Share, Terminal, XlAlignment


def render_canvas(ws, contents_doc):
    """キャンバスの編集
    """
    print("🔧　キャンバスの編集")

    # ウィンドウ枠の固定
    ws.freeze_panes = 'C2'

    # Trellis では、タテ：ヨコ＝３：３ で、１ユニットセルとします。
    # また、上辺、右辺、下辺、左辺に、１セル幅の定規を置きます
    canvas_obj = Canvas.from_dict(contents_doc['canvas'])
    canvas_rect = canvas_obj.rect_obj

    # 横幅または縦幅が１アウト未満の場合は、定規は描画しません
    if canvas_rect.width_obj.total_of_out_counts_qty < 1 or canvas_rect.height_obj.total_of_out_counts_qty < 1:
        return

    # 行の横幅
    for column_th in range(
            canvas_rect.left_obj.total_of_out_counts_th,
            canvas_rect.left_obj.total_of_out_counts_th + canvas_rect.width_obj.total_of_out_counts_qty):
        column_letter = xl.utils.get_column_letter(column_th)
        ws.column_dimensions[column_letter].width = 2.7    # 2.7 characters = about 30 pixels

    # 列の高さ
    for row_th in range(
            canvas_rect.top_obj.total_of_out_counts_th,
            canvas_rect.top_obj.total_of_out_counts_th + canvas_rect.height_obj.total_of_out_counts_qty):
        ws.row_dimensions[row_th].height = 15    # 15 points = about 30 pixels


def render_all_xl_texts(ws, contents_doc):
    """全てのテキストの描画（定規の番号除く）
    """
    print('🔧　全てのテキストの描画')

    # もし、テキストのリストがあれば
    if 'xlTexts' in contents_doc and (xlTexts := contents_doc['xlTexts']):
        for xl_text_dict in xlTexts:

            # テキスト設定
            if 'text' in xl_text_dict and (text := xl_text_dict['text']):

                # 位置
                location_obj = None
                if 'location' in xl_text_dict and (location_dict := xl_text_dict['location']):
                    location_obj = Point.from_dict(location_dict)

                # テキストの位置揃え
                xl_alignment_obj = None
                if 'xlAlignment' in xl_text_dict and (xl_alignment_dict := xl_text_dict['xlAlignment']):
                    xl_alignment_obj = XlAlignment.from_dict(xl_alignment_dict)

                # フォント
                xl_font_obj = None
                if 'xlFont' in xl_text_dict and (xl_font_dict := xl_text_dict['xlFont']):
                    xl_font_obj = XlFont.from_dict(xl_font_dict)

                # テキストを入力する
                print_text(
                        ws=ws,
                        location_obj=location_obj,
                        text=text,
                        xl_alignment_obj=xl_alignment_obj,
                        xl_font_obj=xl_font_obj)


def render_all_rectangles(ws, contents_doc):
    """全ての矩形の描画
    """
    print('🔧　全ての矩形の描画')

    # もし、矩形のリストがあれば
    if 'rectangles' in contents_doc and (rectangles_list := contents_doc['rectangles']):

        for rectangle_dict in rectangles_list:

            rect_obj = None
            if 'bounds' in rectangle_dict and (bounds_dict := rectangle_dict['bounds']):
                rect_obj = Rectangle.from_dict(bounds_dict)

                # セル結合
                if 'mergeCells' in rectangle_dict and (is_merge_cells := rectangle_dict['mergeCells']):
                    if is_merge_cells:
                        column_th = rect_obj.left_obj.total_of_out_counts_th
                        row_th = rect_obj.top_obj.total_of_out_counts_th
                        columns = rect_obj.width_obj.total_of_out_counts_qty
                        rows = rect_obj.height_obj.total_of_out_counts_qty

                        if 0 < columns and 0 < rows and (1 < columns or 1 < rows):
                            column_letter = xl.utils.get_column_letter(column_th)
                            column_letter2 = xl.utils.get_column_letter(column_th + columns - 1)
                            ws.merge_cells(f'{column_letter}{row_th}:{column_letter2}{row_th + rows - 1}')

                if 'color' in rectangle_dict and (bg_color := rectangle_dict['color']):
                    # もし境界線が指定されていれば、描画する
                    if 'xlBorder' in rectangle_dict and (xl_border_dict := rectangle_dict['xlBorder']):
                        draw_xl_border_on_rectangle(
                                ws=ws,
                                xl_border_dict=xl_border_dict,
                                column_th=rect_obj.left_obj.total_of_out_counts_th,
                                row_th=rect_obj.top_obj.total_of_out_counts_th,
                                columns=rect_obj.width_obj.total_of_out_counts_qty,
                                rows=rect_obj.height_obj.total_of_out_counts_qty)

                    # 矩形を塗りつぶす
                    fill_rectangle(
                            ws=ws,
                            column_th=rect_obj.left_obj.total_of_out_counts_th,
                            row_th=rect_obj.top_obj.total_of_out_counts_th,
                            columns=rect_obj.width_obj.total_of_out_counts_qty,
                            rows=rect_obj.height_obj.total_of_out_counts_qty,
                            color=bg_color)


def render_all_pillar_rugs(ws, contents_doc):
    """全ての柱の敷物の描画
    """
    print('🔧　全ての柱の敷物の描画')

    # もし、柱のリストがあれば
    if 'pillars' in contents_doc and (pillars_list := contents_doc['pillars']):

        for pillar_dict in pillars_list:
            pillar_obj = Pillar.from_dict(pillar_dict)

            if 'baseColor' in pillar_dict and (base_color := pillar_dict['baseColor']):
                pillar_rect_obj = pillar_obj.rect_obj

                # 矩形を塗りつぶす
                fill_rectangle(
                        ws=ws,
                        column_th=pillar_rect_obj.left_obj.total_of_out_counts_th,
                        row_th=pillar_rect_obj.top_obj.total_of_out_counts_th,
                        columns=pillar_rect_obj.width_obj.total_of_out_counts_qty,
                        rows=pillar_rect_obj.height_obj.total_of_out_counts_qty,
                        color=base_color)


def render_all_card_shadows(ws, contents_doc):
    """全てのカードの影の描画
    """
    print('🔧　全てのカードの影の描画')

    # もし、柱のリストがあれば
    if 'pillars' in contents_doc and (pillars_list := contents_doc['pillars']):

        for pillar_dict in pillars_list:
            pillar_obj = Pillar.from_dict(pillar_dict)

            # もし、カードの辞書があれば
            if 'cards' in pillar_dict and (card_dict_list := pillar_dict['cards']):

                for card_dict in card_dict_list:
                    card_obj = Card.from_dict(card_dict)

                    if 'shadowColor' in card_dict:
                        card_shadow_color = card_dict['shadowColor']

                        card_rect_obj = card_obj.rect_obj

                        # 端子の影を描く
                        fill_rectangle(
                                ws=ws,
                                column_th=card_rect_obj.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                row_th=card_rect_obj.top_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                columns=card_rect_obj.width_obj.total_of_out_counts_qty,
                                rows=card_rect_obj.height_obj.total_of_out_counts_qty,
                                color=card_shadow_color)


def render_all_terminal_shadows(ws, contents_doc):
    """全ての端子の影の描画
    """
    print('🔧　全ての端子の影の描画')

    # もし、柱のリストがあれば
    if 'pillars' in contents_doc and (pillars_list := contents_doc['pillars']):

        for pillar_dict in pillars_list:
            pillar_obj = Pillar.from_dict(pillar_dict)

            # もし、端子のリストがあれば
            if 'terminals' in pillar_dict and (terminals_list := pillar_dict['terminals']):

                for terminal_dict in terminals_list:
                    terminal_obj = Terminal.from_dict(terminal_dict)
                    terminal_rect_obj = terminal_obj.rect_obj

                    terminal_shadow_color = terminal_dict['shadowColor']

                    # 端子の影を描く
                    fill_rectangle(
                            ws=ws,
                            column_th=terminal_rect_obj.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                            row_th=terminal_rect_obj.top_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                            columns=9,
                            rows=9,
                            color=terminal_shadow_color)


def render_all_line_tape_shadows(ws, contents_doc):
    """全てのラインテープの影の描画
    """
    print('🔧　全てのラインテープの影の描画')

    # もし、ラインテープの配列があれば
    if 'lineTapes' in contents_doc and (line_tape_list := contents_doc['lineTapes']):

        for line_tape_dict in line_tape_list:
            for segment_dict in line_tape_dict['segments']:
                if 'shadowColor' in segment_dict and (line_tape_shadow_color := segment_dict['shadowColor']):
                    segment_rect = Rectangle.from_dict(segment_dict)

                    # 端子の影を描く
                    fill_rectangle(
                            ws=ws,
                            column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                            row_th=segment_rect.top_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                            columns=segment_rect.width_obj.total_of_out_counts_qty,
                            rows=segment_rect.height_obj.total_of_out_counts_qty,
                            color=line_tape_shadow_color)


def render_all_cards(ws, contents_doc):
    """全てのカードの描画
    """
    print('🔧　全てのカードの描画')

    # もし、柱のリストがあれば
    if 'pillars' in contents_doc and (pillars_list := contents_doc['pillars']):

        for pillar_dict in pillars_list:
            pillar_obj = Pillar.from_dict(pillar_dict)

            # 柱と柱の隙間（隙間柱）は無視する
            if 'baseColor' not in pillar_dict or not pillar_dict['baseColor']:
                continue

            base_color = pillar_dict['baseColor']
            card_list = pillar_dict['cards']

            for card_dict in card_list:
                card_obj = Card.from_dict(card_dict)
                card_rect_obj = card_obj.rect_obj

                try:
                    # ヘッダーの矩形の枠線を描きます
                    draw_rectangle(
                            ws=ws,
                            column_th=card_rect_obj.left_obj.total_of_out_counts_th,
                            row_th=card_rect_obj.top_obj.total_of_out_counts_th,
                            columns=card_rect_obj.width_obj.total_of_out_counts_qty,
                            rows=card_rect_obj.height_obj.total_of_out_counts_qty)
                except:
                    print(f'ERROR: render_all_cards: {card_dict=}')
                    raise

                if 'paperStrips' in card_dict:
                    paper_strip_list = card_dict['paperStrips']

                    for index, paper_strip in enumerate(paper_strip_list):

                        # 短冊１行の描画
                        render_paper_strip(
                                ws=ws,
                                paper_strip=paper_strip,
                                column_th=card_rect_obj.left_obj.total_of_out_counts_th,
                                row_th=index * Share.OUT_COUNTS_THAT_CHANGE_INNING + card_rect_obj.top_obj.total_of_out_counts_th,
                                columns=card_rect_obj.width_obj.total_of_out_counts_qty,
                                rows=card_rect_obj.height_obj.total_of_out_counts_qty)


def render_all_terminals(ws, contents_doc):
    """全ての端子の描画
    """
    print('🔧　全ての端子の描画')

    # もし、柱のリストがあれば
    if 'pillars' in contents_doc and (pillars_list := contents_doc['pillars']):

        for pillar_dict in pillars_list:
            pillar_obj = Pillar.from_dict(pillar_dict)

            # もし、端子のリストがあれば
            if 'terminals' in pillar_dict and (terminals_list := pillar_dict['terminals']):

                for terminal_dict in terminals_list:
                    terminal_obj = Terminal.from_dict(terminal_dict)
                    terminal_rect_obj = terminal_obj.rect_obj

                    terminal_pixel_art = terminal_dict['pixelArt']

                    if terminal_pixel_art == 'start':
                        # 始端のドット絵を描く
                        fill_start_terminal(
                            ws=ws,
                            column_th=terminal_rect_obj.left_obj.total_of_out_counts_th,
                            row_th=terminal_rect_obj.top_obj.total_of_out_counts_th)

                    elif terminal_pixel_art == 'end':
                        # 終端のドット絵を描く
                        fill_end_terminal(
                            ws=ws,
                            column_th=terminal_rect_obj.left_obj.total_of_out_counts_th,
                            row_th=terminal_rect_obj.top_obj.total_of_out_counts_th)


def render_all_line_tapes(ws, contents_doc):
    """全てのラインテープの描画
    """
    print('🔧　全てのラインテープの描画')

    # もし、ラインテープの配列があれば
    if 'lineTapes' in contents_doc and (line_tape_list := contents_doc['lineTapes']):

        # 各ラインテープ
        for line_tape_dict in line_tape_list:

            line_tape_outline_color = None
            if 'outlineColor' in line_tape_dict:
                line_tape_outline_color = line_tape_dict['outlineColor']

            # 各セグメント
            for segment_dict in line_tape_dict['segments']:

                line_tape_direction = None
                if 'direction' in segment_dict:
                    line_tape_direction = segment_dict['direction']

                if 'color' in segment_dict:
                    line_tape_color = segment_dict['color']

                    segment_rect = Rectangle.from_dict(segment_dict)

                    # ラインテープを描く
                    fill_rectangle(
                            ws=ws,
                            column_th=segment_rect.left_obj.total_of_out_counts_th,
                            row_th=segment_rect.top_obj.total_of_out_counts_th,
                            columns=segment_rect.width_obj.total_of_out_counts_qty,
                            rows=segment_rect.height_obj.total_of_out_counts_qty,
                            color=line_tape_color)

                    # （あれば）アウトラインを描く
                    if line_tape_outline_color and line_tape_direction:
                        outline_fill_obj = ColorSystem.var_color_name_to_fill_obj(line_tape_outline_color)

                        # （共通処理）垂直方向
                        if line_tape_direction in ['from_here.falling_down', 'after_go_right.turn_falling_down', 'after_go_left.turn_up', 'after_go_left.turn_falling_down']:
                            # 左辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=1,
                                    rows=segment_rect.height_obj.total_of_out_counts_qty - 2,
                                    color=line_tape_outline_color)

                            # 右辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + segment_rect.width_obj.total_of_out_counts_qty,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=1,
                                    rows=segment_rect.height_obj.total_of_out_counts_qty - 2,
                                    color=line_tape_outline_color)

                        # （共通処理）水平方向
                        elif line_tape_direction in ['after_falling_down.turn_right', 'continue.go_right', 'after_falling_down.turn_left', 'continue.go_left', 'after_up.turn_right', 'from_here.go_right']:
                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=segment_rect.width_obj.total_of_out_counts_qty - 2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + segment_rect.height_obj.total_of_out_counts_qty,
                                    columns=segment_rect.width_obj.total_of_out_counts_qty - 2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # ここから落ちていく
                        if line_tape_direction == 'from_here.falling_down':
                            # 左辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th,
                                    columns=1,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 右辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + segment_rect.width_obj.total_of_out_counts_qty,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th,
                                    columns=1,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # 落ちたあと、右折
                        elif line_tape_direction == 'after_falling_down.turn_right':
                            # 左辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=1,
                                    rows=2,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING + 1,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # そのまま右進
                        elif line_tape_direction == 'continue.go_right':
                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # 右進から落ちていく
                        elif line_tape_direction == 'after_go_right.turn_falling_down':
                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 右辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + segment_rect.width_obj.total_of_out_counts_qty,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=1,
                                    rows=2,
                                    color=line_tape_outline_color)

                        # 落ちたあと左折
                        elif line_tape_direction == 'after_falling_down.turn_left':
                            # 右辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + segment_rect.width_obj.total_of_out_counts_qty,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=1,
                                    rows=2,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + segment_rect.width_obj.total_of_out_counts_qty - Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING + 1,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # そのまま左進
                        elif line_tape_direction == 'continue.go_left':
                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=segment_rect.width_obj.total_of_out_counts_qty,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=segment_rect.width_obj.total_of_out_counts_qty,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # 左進から上っていく
                        elif line_tape_direction == 'after_go_left.turn_up':
                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + segment_rect.height_obj.total_of_out_counts_qty,
                                    columns=2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 左辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + segment_rect.height_obj.total_of_out_counts_qty - 2,
                                    columns=1,
                                    rows=3,
                                    color=line_tape_outline_color)

                            # 右辺（横長）を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + segment_rect.height_obj.total_of_out_counts_qty - 2,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # 上がってきて右折
                        elif line_tape_direction == 'after_up.turn_right':
                            # 左辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th,
                                    columns=1,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING + 1,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # 左進から落ちていく
                        elif line_tape_direction == 'after_go_left.turn_falling_down':
                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=2 * Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 左辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th - 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=1,
                                    rows=segment_rect.height_obj.total_of_out_counts_qty,
                                    color=line_tape_outline_color)

                            # 右辺（横長）を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th + Share.OUT_COUNTS_THAT_CHANGE_INNING + 1,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING - 1,
                                    rows=1,
                                    color=line_tape_outline_color)

                        # ここから右進
                        elif line_tape_direction == 'from_here.go_right':
                            # 上辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th - 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)

                            # 下辺を描く
                            fill_rectangle(
                                    ws=ws,
                                    column_th=segment_rect.left_obj.total_of_out_counts_th,
                                    row_th=segment_rect.top_obj.total_of_out_counts_th + 1,
                                    columns=Share.OUT_COUNTS_THAT_CHANGE_INNING,
                                    rows=1,
                                    color=line_tape_outline_color)


def fill_rectangle(ws, column_th, row_th, columns, rows, color):
    """矩形を塗りつぶします
    """
    color_type = ColorSystem.what_is_var_color_name(color)
    if color_type != ColorSystem.DARKNESS:
        fill_obj = ColorSystem.var_color_name_to_fill_obj(color)

    # 横へ
    for cur_column_th in range(column_th, column_th + columns):
        column_letter = xl.utils.get_column_letter(cur_column_th)

        # 縦へ
        for cur_row_th in range(row_th, row_th + rows):
            cell = ws[f'{column_letter}{cur_row_th}']

            if color_type == ColorSystem.DARKNESS:
                # TODO セルの背景色を取得
                # TODO ウェブ・セーフ・カラーに変換
                # TODO さらに影の色に変換
                # TODO 指定によりそれを複数回
                fill_obj = ColorSystem.var_color_name_to_fill_obj(color)
            
            cell.fill = fill_obj


def draw_xl_border_on_rectangle(ws, xl_border_dict, column_th, row_th, columns, rows):
    """境界線の描画
    """
    top_side = None
    right_side = None
    bottom_side = None
    left_side = None

    # 罫線の style の種類
    # 📖 [openpyxl.styles.borders module](https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.styles.borders.html)
    # ‘mediumDashed’, ‘mediumDashDotDot’, ‘dashDot’, ‘dashed’, ‘slantDashDot’, ‘dashDotDot’, ‘thick’, ‘thin’, ‘dotted’, ‘double’, ‘medium’, ‘hair’, ‘mediumDashDot’

    if 'top' in xl_border_dict and (top_dict := xl_border_dict['top']):
        web_safe_color_code = None
        style = None

        if 'color' in top_dict and (color := top_dict['color']):
            web_safe_color_code = ColorSystem.var_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in top_dict and (style := top_dict['xlStyle']):
            pass

        try:
            top_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    if 'right' in xl_border_dict and (right_dict := xl_border_dict['right']):
        web_safe_color_code = None
        style = None

        if 'color' in right_dict and (color := right_dict['color']):
            web_safe_color_code = ColorSystem.var_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in right_dict and (style := right_dict['xlStyle']):
            pass

        try:
            right_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    if 'bottom' in xl_border_dict and (bottom_dict := xl_border_dict['bottom']):
        web_safe_color_code = None
        style = None

        if 'color' in bottom_dict and (color := bottom_dict['color']):
            web_safe_color_code = ColorSystem.var_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in bottom_dict and (style := bottom_dict['xlStyle']):
            pass

        try:
            bottom_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    if 'left' in xl_border_dict and (left_dict := xl_border_dict['left']):
        web_safe_color_code = None
        style = None

        if 'color' in left_dict and (color := left_dict['color']):
            web_safe_color_code = ColorSystem.var_color_name_to_web_safe_color_code(color)

        if 'xlStyle' in left_dict and (style := left_dict['xlStyle']):
            pass

        try:
            left_side = Side(style=style, color=ColorSystem.web_safe_color_code_to_xl(web_safe_color_code))
        except:
            print(f'draw_xl_border_on_rectangle: いずれかが、未対応の指定： {style=} {web_safe_color_code=}')


    # TODO 厚みが１のケースや、角は、２辺に線を引く

    
    top_border = Border(top=top_side)           # 上辺
    right_border = Border(right=right_side)     # 右辺
    bottom_border = Border(bottom=bottom_side)  # 下辺
    left_border = Border(left=left_side)        # 左辺

    # 水平方向
    if rows == 0 or rows == 1:
        if rows == 0:
            # 上辺だけ引く
            horizontal_border = Border(top=top_side)
        else:
            # 上辺と下辺の両方を引く
            horizontal_border = Border(top=top_side, bottom=bottom_side)

        # （角を除く）横へ
        for cur_column_th in range(column_th + 1, column_th + columns - 1):
            column_letter = xl.utils.get_column_letter(cur_column_th)
            cell = ws[f'{column_letter}{row_th}']
            cell.border = horizontal_border

    # 上辺を引くのと、下辺を引くのとがある
    else:
        top_border = Border(top=top_side)
        bottom_border = Border(bottom=bottom_side)

        # （角を除く）横へ
        for cur_column_th in range(column_th + 1, column_th + columns - 1):
            column_letter = xl.utils.get_column_letter(cur_column_th)

            cell = ws[f'{column_letter}{row_th}']
            cell.border = top_border

            cell = ws[f'{column_letter}{row_th + rows - 1}']
            cell.border = bottom_border


    # 垂直方向
    if columns == 0 or columns == 1:
        if columns == 0:
            # 左辺だけ引く
            vertical_border = Border(left=left_side)
        else:
            # 左辺と右辺の両方を引く
            vertical_border = Border(left=left_side, right=right_side)

        # （角を除く）縦へ
        for cur_row_th in range(row_th + 1, row_th + rows - 1):
            column_letter = xl.utils.get_column_letter(columns)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = vertical_border

    # 左辺を引くのと、右辺を引くのとがある
    else:
        left_border = Border(left=left_side)
        right_border = Border(right=right_side)

        # （角を除く）縦へ
        for cur_row_th in range(row_th + 1, row_th + rows - 1):
            column_letter = xl.utils.get_column_letter(column_th)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = left_border

            column_letter = xl.utils.get_column_letter(column_th + columns - 1)
            cell = ws[f'{column_letter}{cur_row_th}']
            cell.border = right_border


    # 左上隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, left=left_side)

    # 右上隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th + columns - 1)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, right=right_side)

    # 左下隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th + rows - 1}']
        cell.border = Border(left=left_side, bottom=bottom_side)

    # 右下隅
    if 1 < columns and 1 < rows:
        column_letter = xl.utils.get_column_letter(column_th + columns - 1)
        cell = ws[f'{column_letter}{row_th + rows - 1}']
        cell.border = Border(right=right_side, bottom=bottom_side)

    # 四方
    if columns == 1 and rows == 1:
        column_letter = xl.utils.get_column_letter(column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.border = Border(top=top_side, right=right_side, bottom=bottom_side, left=left_side)


def print_text(ws, location_obj, text, xl_alignment_obj, xl_font_obj):
    """テキスト描画
    """

    # テキストの位置
    column_th = location_obj.x_obj.total_of_out_counts_th
    row_th = location_obj.y_obj.total_of_out_counts_th

    # テキスト設定
    column_letter = xl.utils.get_column_letter(column_th)
    cell = ws[f'{column_letter}{row_th}']
    cell.value = text

    # フォント設定
    if xl_font_obj:
        cell.font = Font(color=xl_font_obj.color_code_for_xl)

    # テキストの位置揃え
    if xl_alignment_obj:
        cell.alignment = Alignment(
                horizontal=xl_alignment_obj.xlHorizontal,
                vertical=xl_alignment_obj.xlVertical)


def draw_rectangle(ws, column_th, row_th, columns, rows):
    """矩形の枠線の描画
    """

    # 赤はデバッグ用
    red_side = Side(style='thick', color='FF0000')
    black_side = Side(style='thick', color='000000')

    red_top_border = Border(top=red_side)
    red_top_right_border = Border(top=red_side, right=red_side)
    red_right_border = Border(right=red_side)
    red_bottom_right_border = Border(bottom=red_side, right=red_side)
    red_bottom_border = Border(bottom=red_side)
    red_bottom_left_border = Border(bottom=red_side, left=red_side)
    red_left_border = Border(left=red_side)
    red_top_left_border = Border(top=red_side, left=red_side)

    black_top_border = Border(top=black_side)
    black_top_right_border = Border(top=black_side, right=black_side)
    black_right_border = Border(right=black_side)
    black_bottom_right_border = Border(bottom=black_side, right=black_side)
    black_bottom_border = Border(bottom=black_side)
    black_bottom_left_border = Border(bottom=black_side, left=black_side)
    black_left_border = Border(left=black_side)
    black_top_left_border = Border(top=black_side, left=black_side)

    # 罫線で四角を作る　＞　左上
    cur_column_th = column_th
    column_letter = xl.utils.get_column_letter(cur_column_th)
    cur_row_th = row_th
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_top_left_border

    # 罫線で四角を作る　＞　上辺
    for cur_column_th in range(column_th + 1, column_th + columns - 1):
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_top_border

    # 罫線で四角を作る　＞　右上
    cur_column_th = column_th + columns - 1
    column_letter = xl.utils.get_column_letter(cur_column_th)
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_top_right_border

    # 罫線で四角を作る　＞　左辺
    cur_column_th = column_th
    for cur_row_th in range(row_th + 1, row_th + rows - 1):
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_left_border

    # 罫線で四角を作る　＞　左下
    cur_row_th = row_th + rows - 1
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_bottom_left_border

    # 罫線で四角を作る　＞　下辺
    for cur_column_th in range(column_th + 1, column_th + columns - 1):
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_bottom_border

    # 罫線で四角を作る　＞　右下
    cur_column_th = column_th + columns - 1
    column_letter = xl.utils.get_column_letter(cur_column_th)
    cell = ws[f'{column_letter}{cur_row_th}']
    cell.border = black_bottom_right_border

    # 罫線で四角を作る　＞　右辺
    for cur_row_th in range(row_th + 1, row_th + rows - 1):
        cell = ws[f'{column_letter}{cur_row_th}']
        cell.border = black_right_border


def render_paper_strip(ws, paper_strip, column_th, row_th, columns, rows):
    """短冊１行の描画
    """

    # 柱のヘッダーの背景色
    if 'bgColor' in paper_strip and (base_color := paper_strip['bgColor']):
        # 矩形を塗りつぶす
        fill_rectangle(
                ws=ws,
                column_th=column_th,
                row_th=row_th,
                columns=columns,
                rows=1 * Share.OUT_COUNTS_THAT_CHANGE_INNING,   # １行分
                color=base_color)

    # インデント
    if 'indent' in paper_strip:
        indent = paper_strip['indent']
    else:
        indent = 0

    # アイコン（があれば画像をワークシートのセルに挿入）
    if 'icon' in paper_strip:
        image_basename = paper_strip['icon']  # 例： 'white-game-object.png'

        cur_column_th = column_th + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        #
        # NOTE 元の画像サイズで貼り付けられるわけではないの、何でだろう？ 60x60pixels の画像にしておくと、90x90pixels のセルに合う？
        #
        # TODO 📖 [PythonでExcelファイルに画像を挿入する/列の幅を調整する](https://qiita.com/kaba_san/items/b231a41891ebc240efc7)
        # 難しい
        #
        try:
            ws.add_image(XlImage(os.path.join('./assets/icons', image_basename)), f"{column_letter}{row_th}")
        except FileNotFoundError as e:
            print(f'FileNotFoundError {e=} {image_basename=}')


    # テキスト（があれば）
    if 'text0' in paper_strip:
        text = paper_strip['text0']

        # 左に１マス分のアイコンを置く前提
        icon_columns = Share.OUT_COUNTS_THAT_CHANGE_INNING
        cur_column_th = column_th + icon_columns + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{row_th}']
        cell.value = text

    if 'text1' in paper_strip:
        text = paper_strip['text1']

        # 左に１マス分のアイコンを置く前提
        icon_columns = Share.OUT_COUNTS_THAT_CHANGE_INNING
        cur_column_th = column_th + icon_columns + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{row_th + 1}']
        cell.value = text

    if 'text3' in paper_strip:
        text = paper_strip['text2']

        # 左に１マス分のアイコンを置く前提
        icon_columns = Share.OUT_COUNTS_THAT_CHANGE_INNING
        cur_column_th = column_th + icon_columns + (indent * Share.OUT_COUNTS_THAT_CHANGE_INNING)
        column_letter = xl.utils.get_column_letter(cur_column_th)
        cell = ws[f'{column_letter}{row_th + 2}']
        cell.value = text


def fill_start_terminal(ws, column_th, row_th):
    """始端を描きます
    """
    # ドット絵を描きます
    fill_pixel_art(
            ws=ws,
            column_th=column_th,
            row_th=row_th,
            columns=9,
            rows=9,
            pixels=[
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
                [1, 1, 1, 0, 0, 0, 1, 1, 1],
                [1, 0, 0, 1, 1, 1, 0, 0, 1],
                [1, 1, 0, 1, 1, 1, 1, 0, 1],
                [1, 1, 1, 0, 0, 0, 1, 1, 1],
                [1, 0, 1, 1, 1, 1, 0, 1, 1],
                [1, 0, 0, 1, 1, 1, 0, 0, 1],
                [1, 1, 1, 0, 0, 0, 1, 1, 1],
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
            ])


def fill_end_terminal(ws, column_th, row_th):
    """終端を描きます
    """
    # ドット絵を描きます
    fill_pixel_art(
            ws=ws,
            column_th=column_th,
            row_th=row_th,
            columns=9,
            rows=9,
            pixels=[
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 0, 0, 0, 0, 0, 0, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 0, 0, 0, 0, 0, 0, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 1, 1, 1, 1, 1, 1, 1],
                [1, 0, 0, 0, 0, 0, 0, 0, 1],
                [1, 1, 1, 1, 1, 1, 1, 1, 1],
            ])


def fill_pixel_art(ws, column_th, row_th, columns, rows, pixels):
    """ドット絵を描きます
    """
    # 背景色
    mat_black = PatternFill(patternType='solid', fgColor='080808')
    mat_white = PatternFill(patternType='solid', fgColor='E8E8E8')

    # 横へ
    for cur_column_th in range(column_th, column_th + columns):
        for cur_row_th in range(row_th, row_th + rows):
            column_letter = xl.utils.get_column_letter(cur_column_th)
            cell = ws[f'{column_letter}{cur_row_th}']

            pixel = pixels[cur_row_th - row_th][cur_column_th - column_th]
            if pixel == 1:
                cell.fill = mat_black
            else:
                cell.fill = mat_white
